#!/usr/bin/env python3
"""Audit whether GWAS candidate genes are present in bulk expression datasets."""

from __future__ import annotations

import csv
import gzip
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW = PROJECT_ROOT / "data" / "raw_downloads"
GWAS = PROJECT_ROOT / "results" / "gwas"
OUT_DIR = PROJECT_ROOT / "results" / "bulk"
OUT_TSV = OUT_DIR / "candidate_gene_bulk_presence.tsv"
OUT_MD = OUT_DIR / "candidate_gene_bulk_presence_summary.md"

CANDIDATES = GWAS / "gwas_candidate_gene_universe.tsv"
GSE234354 = RAW / "GSE234354__GSE234354_gene_count_matrix.txt.gz"
GSE313775 = RAW / "GSE313775__GSE313775_rawCountMatrix.tsv.gz"
GSE141549 = RAW / "GSE141549__GSE141549_batchCorrectednormalizedArrayscombined.xlsx"
GSE51981 = RAW / "GSE51981__GSE51981_series_matrix.txt.gz"

HIGHLIGHT_GENES = ["FSHB", "WNT4", "KDR", "ESR1", "SSPN", "ITPR2", "LY96", "HRH1"]


def read_candidates() -> list[dict[str, str]]:
    with CANDIDATES.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_first_col_gzip(path: Path) -> set[str]:
    values: set[str] = set()
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        next(reader, None)
        for row in reader:
            if row:
                values.add(row[0].strip())
    return values


def read_gse313775_ids_symbols(path: Path) -> tuple[set[str], set[str]]:
    gene_ids: set[str] = set()
    symbols: set[str] = set()
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            gene_id = (row.get("Gene") or "").strip()
            symbol = (row.get("Symbol") or "").strip()
            if gene_id:
                gene_ids.add(gene_id)
            if symbol:
                symbols.add(symbol)
    return gene_ids, symbols


def read_gse141549_symbols(path: Path) -> Counter[str]:
    counts: Counter[str] = Counter()
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(values_only=True))]
    try:
        symbol_index = header.index("Gene_symbol")
    except ValueError as exc:
        raise ValueError("GSE141549 workbook is missing Gene_symbol column") from exc
    for row in worksheet.iter_rows(values_only=True):
        symbol = row[symbol_index]
        if symbol:
            counts[str(symbol).strip()] += 1
    return counts


def read_gse51981_probe_ids(path: Path) -> set[str]:
    probe_ids: set[str] = set()
    in_table = False
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for line in handle:
            line = line.rstrip("\n")
            if line == "!series_matrix_table_begin":
                in_table = True
                continue
            if line == "!series_matrix_table_end":
                break
            if not in_table:
                continue
            row = next(csv.reader([line], delimiter="\t"))
            if row and row[0].strip().strip('"') != "ID_REF":
                probe_ids.add(row[0].strip().strip('"'))
    return probe_ids


def yes_no(value: bool) -> str:
    return "yes" if value else "no"


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    candidates = read_candidates()

    gse234354_gene_ids = read_first_col_gzip(GSE234354)
    gse313775_gene_ids, gse313775_symbols = read_gse313775_ids_symbols(GSE313775)
    gse141549_symbol_counts = read_gse141549_symbols(GSE141549)
    gse51981_probe_ids = read_gse51981_probe_ids(GSE51981)

    rows: list[dict[str, str]] = []
    for candidate in candidates:
        gene_symbol = candidate["gene_symbol"]
        gene_id = candidate["gene_id"]
        rows.append(
            {
                "gene_symbol": gene_symbol,
                "gene_id": gene_id,
                "genetic_priority": candidate["genetic_priority"],
                "neighborhood_id": candidate["neighborhood_id"],
                "ld_neighborhood_class": candidate["ld_neighborhood_class"],
                "module_hint_preliminary": candidate["module_hint_preliminary"],
                "GSE234354_present_by_ensembl_id": yes_no(gene_id in gse234354_gene_ids),
                "GSE313775_present_by_ensembl_id": yes_no(gene_id in gse313775_gene_ids),
                "GSE313775_present_by_symbol": yes_no(bool(gene_symbol) and gene_symbol in gse313775_symbols),
                "GSE141549_present_by_symbol": yes_no(bool(gene_symbol) and gene_symbol in gse141549_symbol_counts),
                "GSE141549_probe_count_for_symbol": str(gse141549_symbol_counts.get(gene_symbol, 0)) if gene_symbol else "0",
                "GSE51981_probe_mapping_status": "requires_GPL570_annotation",
                "GSE51981_raw_probe_id_match": yes_no(gene_id in gse51981_probe_ids or gene_symbol in gse51981_probe_ids),
            }
        )

    fieldnames = list(rows[0].keys())
    with OUT_TSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    def count_present(field: str) -> int:
        return sum(row[field] == "yes" for row in rows)

    lines = [
        "# Candidate Gene Bulk Presence Audit",
        "",
        "## Scope",
        "",
        "This audit checks whether the GWAS-derived candidate genes can be matched to available bulk expression matrices before any expression-effect testing.",
        "",
        "## Dataset Match Counts",
        "",
        f"- Candidate gene records: {len(rows):,}",
        f"- GSE234354 present by Ensembl ID: {count_present('GSE234354_present_by_ensembl_id'):,}",
        f"- GSE313775 present by Ensembl ID: {count_present('GSE313775_present_by_ensembl_id'):,}",
        f"- GSE313775 present by gene symbol: {count_present('GSE313775_present_by_symbol'):,}",
        f"- GSE141549 present by gene symbol: {count_present('GSE141549_present_by_symbol'):,}",
        "- GSE51981 requires GPL570 probe-to-gene annotation before gene-level matching.",
        "",
        "## Highlight Genes",
        "",
        "| Gene | GSE234354 | GSE313775 ID | GSE313775 symbol | GSE141549 symbol/probes | GSE51981 |",
        "|---|---|---|---|---|---|",
    ]
    by_symbol = {row["gene_symbol"]: row for row in rows if row["gene_symbol"]}
    for gene in HIGHLIGHT_GENES:
        row = by_symbol.get(gene)
        if not row:
            lines.append(f"| {gene} | not_in_candidate_universe | not_in_candidate_universe | not_in_candidate_universe | not_in_candidate_universe | requires_GPL570_annotation |")
            continue
        gse141549 = f"{row['GSE141549_present_by_symbol']}/{row['GSE141549_probe_count_for_symbol']}"
        lines.append(
            f"| {gene} | {row['GSE234354_present_by_ensembl_id']} | {row['GSE313775_present_by_ensembl_id']} | {row['GSE313775_present_by_symbol']} | {gse141549} | {row['GSE51981_probe_mapping_status']} |"
        )

    lines.extend(
        [
            "",
            "## Interpretation Gate",
            "",
            "GSE234354, GSE313775 and GSE141549 can proceed to candidate-level expression extraction. GSE51981 must first obtain GPL570 probe annotation or another reliable probe-to-symbol mapping before being used for candidate gene validation.",
            "",
            "## Output",
            "",
            f"- Presence table: `{OUT_TSV.relative_to(PROJECT_ROOT)}`",
            "",
        ]
    )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_TSV}")
    print(f"Wrote {OUT_MD}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
