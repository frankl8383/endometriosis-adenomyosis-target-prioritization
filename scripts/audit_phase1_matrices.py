#!/usr/bin/env python3
"""Audit downloaded phase 1 expression matrices and metadata tables.

This script checks structural readiness only: file readability, dimensions,
sample-column counts, and metadata hints. It does not run differential
expression or biological interpretation.
"""

from __future__ import annotations

import csv
import gzip
import re
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW = PROJECT_ROOT / "data" / "raw_downloads"
OUT_TSV = PROJECT_ROOT / "results" / "phase1_matrix_audit.tsv"
OUT_MD = PROJECT_ROOT / "results" / "phase1_matrix_audit.md"


def audit_delimited_gzip(label: str, path: Path) -> dict[str, str]:
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        header = next(reader)
        n_rows = 0
        first_data = ""
        for row in reader:
            if not first_data and row:
                first_data = "|".join(row[:6])
            n_rows += 1
    return {
        "dataset": label,
        "file": path.name,
        "file_type": "gzipped_tabular_matrix",
        "n_rows": str(n_rows),
        "n_columns": str(len(header)),
        "sample_or_data_columns": str(max(len(header) - 1, 0)),
        "sheet_or_table": "",
        "header_preview": "|".join(header[:10]),
        "first_data_preview": first_data[:500],
        "status": "ready" if n_rows > 1000 and len(header) > 10 else "review",
    }


def audit_geo_series_matrix(label: str, path: Path) -> dict[str, str]:
    sample_accessions: list[str] = []
    sample_titles: list[str] = []
    characteristic_keys: set[str] = set()
    in_table = False
    table_header: list[str] = []
    table_rows = 0

    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.rstrip("\n")
            if line.startswith("!Sample_geo_accession"):
                sample_accessions.extend(re.findall(r'"([^"]+)"', line))
            elif line.startswith("!Sample_title"):
                sample_titles.extend(re.findall(r'"([^"]+)"', line))
            elif line.startswith("!Sample_characteristics_ch1"):
                for value in re.findall(r'"([^"]+)"', line):
                    if ":" in value:
                        characteristic_keys.add(value.split(":", 1)[0].strip())
            elif line == "!series_matrix_table_begin":
                in_table = True
            elif line == "!series_matrix_table_end":
                in_table = False
            elif in_table and not table_header:
                table_header = next(csv.reader([line], delimiter="\t"))
            elif in_table:
                table_rows += 1

    sample_count = len(sample_accessions)
    table_columns = len(table_header)
    return {
        "dataset": label,
        "file": path.name,
        "file_type": "geo_series_matrix",
        "n_rows": str(table_rows),
        "n_columns": str(table_columns),
        "sample_or_data_columns": str(max(table_columns - 1, sample_count)),
        "sheet_or_table": "series_matrix_table",
        "header_preview": "|".join(table_header[:10]),
        "first_data_preview": "; ".join(sorted(characteristic_keys))[:500],
        "status": "ready" if sample_count >= 100 and table_rows > 1000 else "review",
    }


def audit_xlsx(label: str, path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        second_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
        n_rows = sheet.max_row or 0
        n_cols = sheet.max_column or 0
        rows.append(
            {
                "dataset": label,
                "file": path.name,
                "file_type": "xlsx_workbook",
                "n_rows": str(n_rows),
                "n_columns": str(n_cols),
                "sample_or_data_columns": str(max(n_cols - 1, 0)),
                "sheet_or_table": sheet_name,
                "header_preview": "|".join("" if value is None else str(value) for value in first_row[:10]),
                "first_data_preview": "|".join("" if value is None else str(value) for value in second_row[:10])[:500],
                "status": "ready" if n_rows > 1 and n_cols > 1 else "review",
            }
        )
    workbook.close()
    return rows


def write_markdown(rows: list[dict[str, str]]) -> None:
    ready = sum(row["status"] == "ready" for row in rows)
    lines = [
        "# Phase 1 Matrix Structure Audit",
        "",
        "## Summary",
        "",
        f"- Ready tables/sheets: {ready}/{len(rows)}",
        "",
        "## Tables",
        "",
        "| Dataset | File | Sheet/Table | Rows | Columns | Status | Header Preview |",
        "|---|---|---|---:|---:|---|---|",
    ]
    for row in rows:
        lines.append(
            "| {dataset} | `{file}` | {sheet_or_table} | {n_rows} | {n_columns} | {status} | `{header_preview}` |".format(
                **row
            )
        )
    lines.extend(
        [
            "",
            "## Strict Review Gate",
            "",
            "The downloaded bulk/immune resources pass Phase 1 matrix readiness only if all core files are readable and have plausible row/column dimensions. Passing this audit does not imply that sample metadata are adequate for every planned contrast; metadata harmonization is the next required gate.",
            "",
        ]
    )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    rows: list[dict[str, str]] = []
    rows.append(
        audit_delimited_gzip(
            "GSE234354",
            RAW / "GSE234354__GSE234354_gene_count_matrix.txt.gz",
        )
    )
    rows.append(
        audit_delimited_gzip(
            "GSE313775",
            RAW / "GSE313775__GSE313775_rawCountMatrix.tsv.gz",
        )
    )
    rows.append(
        audit_geo_series_matrix(
            "GSE51981",
            RAW / "GSE51981__GSE51981_series_matrix.txt.gz",
        )
    )
    rows.extend(
        audit_xlsx(
            "GSE141549_sample_link",
            RAW / "GSE141549__GSE141549_Sample_link.xlsx",
        )
    )
    rows.extend(
        audit_xlsx(
            "GSE141549_expression",
            RAW / "GSE141549__GSE141549_batchCorrectednormalizedArrayscombined.xlsx",
        )
    )

    OUT_TSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_TSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)
    write_markdown(rows)
    print(f"Wrote {OUT_TSV}")
    print(f"Wrote {OUT_MD}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
